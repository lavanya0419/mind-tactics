import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter

def generate_excel_report(results, start_time, total_duration, output_file="test_report.xlsx"):
    """
    Generates a beautifully styled 12-sheet Excel workbook from Appium test results.
    """
    total_tests = len(results)
    passed_tests = sum(1 for r in results if r["status"] == "Passed")
    failed_tests = sum(1 for r in results if r["status"] == "Failed")
    skipped_tests = sum(1 for r in results if r["status"] == "Skipped")
    pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0.0

    wb = Workbook()
    
    # --- Sheet 1: Overall Summary ---
    ws_summary = wb.active
    ws_summary.title = "Overall Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Styles
    HEADER_FILL = PatternFill(start_color="0B132B", end_color="0B132B", fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color="1C2541", end_color="1C2541", fill_type="solid")
    KPI_LABEL_FILL = PatternFill(start_color="3A506B", end_color="3A506B", fill_type="solid")
    KPI_VAL_FILL = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
    
    PASS_FILL = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
    FAIL_FILL = PatternFill(start_color="FAD2E1", end_color="FAD2E1", fill_type="solid")
    SKIP_FILL = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
    
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="0B132B")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_kpi_label = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_kpi_val = Font(name="Segoe UI", size=12, bold=True, color="000000")
    font_pass = Font(name="Segoe UI", size=10, bold=True, color="1B4332")
    font_fail = Font(name="Segoe UI", size=10, bold=True, color="7209B7")
    font_regular = Font(name="Segoe UI", size=10, color="000000")
    font_hyperlink = Font(name="Segoe UI", size=10, underline="single", color="0056B3")

    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    double_bottom_border = Border(
        bottom=Side(style='double', color='0B132B'),
        top=Side(style='thin', color='CCCCCC')
    )

    # Title Banner on Summary
    ws_summary.merge_cells("A1:G2")
    title_cell = ws_summary["A1"]
    title_cell.value = "MINDTACTICS AUTOMATION - MOBILE TEST EXECUTIVE DASHBOARD"
    title_cell.font = font_title
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(1, 3):
        for col in range(1, 8):
            cell = ws_summary.cell(row=row, column=col)
            cell.fill = HEADER_FILL

    # KPI Block
    ws_summary["A4"] = "EXECUTIVE SUMMARY"
    ws_summary["A4"].font = font_section
    
    kpi_definitions = [
        ("B5", "B6", "Total Executed", total_tests),
        ("C5", "C6", "Passed", passed_tests),
        ("D5", "D6", "Failed", failed_tests),
        ("E5", "E6", "Skipped", skipped_tests),
        ("F5", "F6", "Pass Rate", f"{pass_rate:.1f}%"),
        ("G5", "G6", "Duration (s)", f"{total_duration:.2f}s")
    ]

    for label_cell, val_cell, label, val in kpi_definitions:
        ws_summary[label_cell] = label
        ws_summary[label_cell].font = font_kpi_label
        ws_summary[label_cell].fill = KPI_LABEL_FILL
        ws_summary[label_cell].alignment = Alignment(horizontal="center", vertical="center")
        ws_summary[label_cell].border = thin_border
        
        ws_summary[val_cell] = val
        ws_summary[val_cell].font = font_kpi_val
        ws_summary[val_cell].fill = KPI_VAL_FILL
        ws_summary[val_cell].alignment = Alignment(horizontal="center", vertical="center")
        ws_summary[val_cell].border = thin_border

    ws_summary["B8"] = "Execution Start Time:"
    ws_summary["B8"].font = Font(name="Segoe UI", size=9, italic=True)
    ws_summary["C8"] = start_time
    ws_summary["C8"].font = Font(name="Segoe UI", size=9)
    ws_summary["C8"].alignment = Alignment(horizontal="left")

    # Add Pie Chart to Summary Sheet at B10
    try:
        chart = PieChart()
        chart.title = "Test Status Distribution"
        
        labels = Reference(ws_summary, min_col=3, max_col=5, min_row=5, max_row=5)
        data = Reference(ws_summary, min_col=3, max_col=5, min_row=6, max_row=6)
        
        chart.add_data(data, from_rows=True)
        chart.set_categories(labels)
        
        from openpyxl.chart.series import DataPoint
        colors = ["D8F3DC", "FAD2E1", "FFF3B0"]
        serie = chart.series[0]
        for i, color in enumerate(colors):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            serie.dPt.append(pt)
            
        ws_summary.add_chart(chart, "B10")
        chart.width = 13
        chart.height = 7.5
    except Exception as chart_err:
        print(f"Skipping chart addition due to: {chart_err}")

    # Set Column widths on Summary
    for col in ws_summary.columns:
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = 15
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["G"].width = 16

    # --- Create 11 Category Sheets ---
    categories = [
        "Functional Testing",
        "UI-UX Testing",
        "Compatibility Testing",
        "Performance Testing",
        "Security Testing",
        "API Testing",
        "Database Testing",
        "Accessibility Testing",
        "Mobile-Specific Testing",
        "Regression Testing",
        "E2E Testing"
    ]

    for cat in categories:
        # Filter results for this category
        cat_results = [r for r in results if r.get("category", "Functional Testing") == cat]
        
        ws_cat = wb.create_sheet(title=cat)
        ws_cat.views.sheetView[0].showGridLines = True
        
        # Section Header
        ws_cat.cell(row=1, column=1, value=cat.upper() + " LOG").font = font_section
        
        # Headers
        headers = ["Index", "Test Name", "Description", "Status", "Duration (s)", "Error Details", "Screenshot Link"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_cat.cell(row=3, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = SUBHEADER_FILL
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 4, 5, 7] else "left", vertical="center")
            cell.border = thin_border

        # Populate rows
        current_row = 4
        for idx, test in enumerate(cat_results, 1):
            cell_idx = ws_cat.cell(row=current_row, column=1, value=idx)
            cell_idx.alignment = Alignment(horizontal="center")
            
            ws_cat.cell(row=current_row, column=2, value=test["name"])
            ws_cat.cell(row=current_row, column=3, value=test["description"])
            
            cell_status = ws_cat.cell(row=current_row, column=4, value=test["status"])
            cell_status.alignment = Alignment(horizontal="center")
            if test["status"] == "Passed":
                cell_status.fill = PASS_FILL
                cell_status.font = font_pass
            elif test["status"] == "Failed":
                cell_status.fill = FAIL_FILL
                cell_status.font = font_fail
            else:
                cell_status.fill = SKIP_FILL
                cell_status.font = font_regular
                
            cell_dur = ws_cat.cell(row=current_row, column=5, value=test["duration"])
            cell_dur.alignment = Alignment(horizontal="center")
            
            ws_cat.cell(row=current_row, column=6, value=test["error_msg"])
            
            cell_screenshot = ws_cat.cell(row=current_row, column=7)
            cell_screenshot.alignment = Alignment(horizontal="center")
            if test["screenshot"]:
                filename = os.path.basename(test["screenshot"])
                relative_link = f"screenshots/{filename}"
                cell_screenshot.value = f'=HYPERLINK("{relative_link}", "View Screenshot")'
                cell_screenshot.font = font_hyperlink
            else:
                cell_screenshot.value = "N/A"
                cell_screenshot.font = font_regular

            for col_idx in range(1, 8):
                cell = ws_cat.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                if col_idx != 4 and col_idx != 7:
                    cell.font = font_regular
                    
            current_row += 1

        # Summary Row
        ws_cat.cell(row=current_row, column=2, value="Category Summary").font = Font(name="Segoe UI", size=10, bold=True)
        cell_total_dur = ws_cat.cell(row=current_row, column=5, value=f"=SUM(E4:E{current_row-1})")
        cell_total_dur.font = Font(name="Segoe UI", size=10, bold=True)
        cell_total_dur.alignment = Alignment(horizontal="center")
        
        for col_idx in range(1, 8):
            cell = ws_cat.cell(row=current_row, column=col_idx)
            cell.border = double_bottom_border

        # Widths
        for col in ws_cat.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if val_str.startswith('='):
                    val_str = "View Screenshot"
                if cell.row in [1, 2]:
                    continue
                max_len = max(max_len, len(val_str))
            ws_cat.column_dimensions[col_letter].width = max(max_len + 3, 11)

        ws_cat.column_dimensions["A"].width = 7
        ws_cat.column_dimensions["B"].width = 25
        ws_cat.column_dimensions["C"].width = 32
        ws_cat.column_dimensions["D"].width = 12
        ws_cat.column_dimensions["E"].width = 14
        ws_cat.column_dimensions["F"].width = 45
        ws_cat.column_dimensions["G"].width = 18

    # Save resiliently
    try:
        wb.save(output_file)
        print(f"Excel report generated successfully at: {os.path.abspath(output_file)}")
    except PermissionError:
        print(f"\n[WARNING] Permission denied to write to '{output_file}'.")
        print("It seems the file is currently open in Excel, WPS Office, or another editor.")
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(output_file)
        fallback_file = f"{base}_{timestamp}{ext}"
        print(f"Saving report to fallback file instead: {fallback_file}\n")
        try:
            wb.save(fallback_file)
            print(f"Excel report generated successfully at: {os.path.abspath(fallback_file)}")
        except Exception as e:
            print(f"Failed to save fallback file: {e}")
